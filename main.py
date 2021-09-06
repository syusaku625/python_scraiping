from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import openpyxl as excel
import pprint

driver = webdriver.Chrome('C:\chromedriver_win32\chromedriver')

driver.get('https://www.navitime.co.jp/category/0503001001/27/?page=7')

wb = excel.Workbook()
sheet = wb.active
sheet.title = 'test_sheet_1'
count=1
for elem in driver.find_elements_by_css_selector("div#spot-list.t_left ul li div.spot-text dl dt.spot-name"):
    print(elem.text)
    sheet.cell(row=count,column=1).value = elem.text
    count=count+1

count=1
for elem in driver.find_elements_by_css_selector("div#spot-list.t_left ul li div.spot-text dl dd dl.spot-detail-section"):
    print(elem.text)
    sheet.cell(row=count,column=2).value = elem.text
    count=count+1

wb.save('page7.xlsx')