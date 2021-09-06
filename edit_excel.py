import openpyxl
wb=openpyxl.load_workbook("test.xlsx")
sheet = wb['test_sheet_1']
maxRow = sheet.max_row + 1

wb2 = openpyxl.Workbook()
sheet2 = wb2.active
sheet2.title = 'test'

count=1
for i in range(1,maxRow):
    #print(sheet.cell(row=i,column=2).value)
    hospital=sheet.cell(row=i,column=1).value
    a=sheet.cell(row=i,column=2).value.split('住所')
    b=a[1].split('電話')
    c=b[0].split('診療')
    d=c[0].split('\n')
    e=d[1].split('府')
    f=e[1].split('市')

    sheet2.cell(row=count,column=1).value = count
    sheet2.cell(row=count,column=2).value = hospital
    sheet2.cell(row=count,column=3).value = e[0]+'府'
    sheet2.cell(row=count,column=4).value = f[0]+'市'
    sheet2.cell(row=count,column=5).value = f[1]
    count=count+1
wb2.save('data.xlsx')



